import openpyxl
class ExcelFile:

    def readExcel(self, path):
        excelFile = openpyxl.load_workbook(path)
        sheet = excelFile.active
        row = 1
        col = 1
        print(sheet.cell(row=1,column = col).value)
        print("Max Row = ")
        print(sheet.max_row)
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=1, column=i).value == 'price':
                col=i
                break

        for i in range(1, sheet.max_row+1):
            for j in range(1, sheet.max_column+1):
                if sheet.cell(row=i,column = j).value == 'Apple':
                    row = i
                    print(row,col)
                    break

        sheet.cell(row,col).value = 99999
        excelFile.save(path)
